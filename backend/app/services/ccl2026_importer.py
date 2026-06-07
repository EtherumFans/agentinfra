"""ccl2026_importer.py — Convert CCL 2025/2026 train.xlsx into iCoDer gold cases.

The CCL (Chinese Computational Linguistics) workshop runs an annual ICD
auto-coding shared task. The 2026 edition uses ~1,800 train rows +
~400 validation rows, sourced from desensitized Chinese EMRs across 10
科室. The gold labels are:
  - col 15  主要疾病诊断 (primary ICD-10, single code)
  - col 16  其他疾病诊断 (other ICD-10, semicolon-separated)
  - col 17  主要手术操作 (primary ICD-9-CM-3, single code or empty)
  - col 18  其他手术操作 (other ICD-9-CM-3, semicolon-separated)

The xlsx lives at ``<ICODER_DATA_ASSET_DIR>/../train.xlsx`` (the iCoDerA
mirror places it at ``E:/iCoDerA/data/train.xlsx``). The asset directory
is read-only; we only write the parsed fixture JSON to
``backend/tests/fixtures/ccl2026_train_gold.json``.

Usage:
    python -m app.services.ccl2026_importer \\
        --xlsx E:/iCoDerA/data/train.xlsx \\
        --out tests/fixtures/ccl2026_train_gold.json
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
logger = logging.getLogger("ccl2026_importer")

# xlsx column layout (1-indexed) — keep in sync with the actual workbook.
COL_ENCOUNTER_ID = 0
COL_CHIEF_COMPLAINT = 1
COL_PRESENT_ILLNESS = 2
COL_PAST_HISTORY = 3
COL_PERSONAL_HISTORY = 4
COL_MARRIAGE_HISTORY = 5
COL_FAMILY_HISTORY = 6
COL_ADMISSION_CONDITION = 7
COL_ADMISSION_DIAGNOSIS = 8
COL_TREATMENT_COURSE = 9
COL_DISCHARGE_CONDITION = 10
COL_DISCHARGE_INSTRUCTIONS = 11
COL_OPERATION_RECORD = 12
COL_PREOPERATIVE_DIAGNOSIS = 13
COL_POSTOPERATIVE_DIAGNOSIS = 14
COL_PRIMARY_DX_CODE = 15
COL_SECONDARY_DX_CODES = 16
COL_PRIMARY_OP_CODE = 17
COL_OTHER_OP_CODES = 18

ICD10_RE = re.compile(r"^[A-Z]\d{2}(\.[\dxX]+\d*)?$")
ICD9_RE = re.compile(r"^\d{2}\.\d{2}[xX]?\d*$")


def _parse_codes(raw) -> list[str]:
    """Semicolon-separated codes → list of normalized non-empty codes."""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s or s.lower() in ("none", "null", "nan", ""):
        return []
    out: list[str] = []
    for token in re.split(r"[;；\s]+", s):
        t = token.strip().strip('"').strip("'")
        if t and t.lower() not in ("none", "null", "nan", ""):
            out.append(t)
    return out


def _normalize_dx(code: str) -> str | None:
    c = code.strip().strip('"').strip("'")
    if not c or not ICD10_RE.match(c):
        return None
    return c


def _normalize_op(code: str) -> str | None:
    c = code.strip().strip('"').strip("'")
    if not c or not ICD9_RE.match(c):
        return None
    return c


def _build_text_blob(row: tuple) -> str:
    """Concatenate the free-text fields into the single text the LLM sees."""
    fields = [
        row[COL_CHIEF_COMPLAINT],
        row[COL_PRESENT_ILLNESS],
        row[COL_PAST_HISTORY],
        row[COL_PERSONAL_HISTORY],
        row[COL_FAMILY_HISTORY],
        row[COL_ADMISSION_CONDITION],
        row[COL_ADMISSION_DIAGNOSIS],
        row[COL_TREATMENT_COURSE],
        row[COL_DISCHARGE_CONDITION],
        row[COL_DISCHARGE_INSTRUCTIONS],
        row[COL_OPERATION_RECORD],
        row[COL_PREOPERATIVE_DIAGNOSIS],
        row[COL_POSTOPERATIVE_DIAGNOSIS],
    ]
    parts: list[str] = []
    for f in fields:
        if f is None:
            continue
        s = str(f).strip()
        if s:
            parts.append(s)
    return "\n".join(parts)


def convert_xlsx(xlsx_path: str | Path) -> list[dict]:
    """Read the CCL xlsx and return a list of iCoDer-shaped gold cases."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb["Sheet1"]
    data_rows = list(ws.iter_rows(values_only=True))
    if not data_rows:
        return []
    # Header is row 0; row 1 in the 2025 workbook is an empty repeat —
    # skip any row whose col 0 is the header label.
    header_label = str(data_rows[0][0]) if data_rows[0][0] is not None else ""
    body = [r for r in data_rows[1:] if r and str(r[0] or "").strip() != header_label]

    out: list[dict] = []
    skipped: list[tuple[int, str]] = []
    for idx, row in enumerate(body):
        try:
            eid = str(row[COL_ENCOUNTER_ID] or "").strip()
            if not eid:
                continue
            primary_dx = _normalize_dx(str(row[COL_PRIMARY_DX_CODE] or ""))
            if not primary_dx:
                skipped.append((idx, "missing primary dx"))
                continue
            secondary_dx = [
                c for c in (_normalize_dx(t) for t in _parse_codes(row[COL_SECONDARY_DX_CODES]))
                if c
            ]
            primary_op_raw = str(row[COL_PRIMARY_OP_CODE] or "").strip()
            primary_op = _normalize_op(primary_op_raw) if primary_op_raw else None
            other_ops = [
                c for c in (_normalize_op(t) for t in _parse_codes(row[COL_OTHER_OP_CODES]))
                if c
            ]
            text = _build_text_blob(row)

            out.append({
                "encounter_id": eid,
                "department": "",  # not provided in this workbook
                "diagnosis_group": "",
                "specialty": "",
                "difficulty": "medium",
                "risk_tags": [],
                "admission_reason": str(row[COL_CHIEF_COMPLAINT] or "").strip(),
                "expected_principal_diagnosis": primary_dx,
                "expected_principal_diag_name": "",  # filled in by loader at eval time
                "expected_principal_procedure": primary_op,
                "expected_principal_proc_name": "",
                "expected_secondary_diagnoses": secondary_dx,
                "expected_procedure_codes": other_ops,
                "expected_drg_group": None,
                "acceptable_alternatives": [],
                "reasoning_expectations": [],
                "evidence_spans": [],
                # CCL-specific extras (kept alongside for downstream use):
                "text": text,
                "source": "CCL2026",
            })
        except Exception as e:
            skipped.append((idx, str(e)))
    logger.info(
        "Converted %d cases from %s (skipped %d)", len(out), xlsx_path, len(skipped),
    )
    if skipped:
        for i, reason in skipped[:5]:
            logger.warning("  skip row %d: %s", i, reason)
    return out


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        default=r"E:\iCoDerA\data\train.xlsx",
        help="Path to the CCL train.xlsx (read-only).",
    )
    parser.add_argument(
        "--out",
        default="tests/fixtures/ccl2026_train_gold.json",
        help="Output JSON path for the gold fixture.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="If > 0, only convert the first N cases (smoke test).",
    )
    args = parser.parse_args(argv)

    cases = convert_xlsx(args.xlsx)
    if args.limit and args.limit > 0:
        cases = cases[: args.limit]

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(cases, f, ensure_ascii=False, indent=2)
    logger.info("Wrote %d cases to %s", len(cases), out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
